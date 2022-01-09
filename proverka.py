import openpyxl
import datetime

def process(fl,num,id, fio,start_ostatok,start_date,kol_dney,end_ostatok,end_date,conclusion,start_row,end_row,koef, url):
    book= openpyxl.load_workbook(filename=fl, data_only=True)
    sheet = book.worksheets[num]
    structure=[]
    list_peremenyh=[]
    lists=[]
    count=0
    #-----------------------------!указать диапазон!------------------------------------
    min_row=start_row
    max_row=end_row
    standart_chislo=koef
    #-----------------------------!указать диапазон!------------------------------------
    x=0
    starting_points=0
    #------------подсчет структур--------------------------------------------------------
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=1, values_only=True):
        for cell in row:
            structure.append(cell)
    N=len(set(structure))
    #------------подсчет структур--------------------------------------------------------
    #------------ подсчет каждого элемента в структуре----------------------------
    podshet = {i:structure.count(i) for i in structure}
    #------------ подсчет каждого элемента в структуре----------------------------

    #------------Создание словаря--------------------------------------------------------

    d={}
    for i in range(min_row, max_row+1):
        key1=sheet[id+str(i)].value
        value1 = [sheet[fio+str(i)].value,sheet[start_ostatok+str(i)].value,sheet[start_date+str(i)].value, sheet[kol_dney+str(i)].value]
        d.setdefault(key1, []).append(value1)
    print(d)
    #------------Создание словаря--------------------------------------------------------


    for key, value in d.items():
        if len(value) > 1:
            for p in range(0, len(value)):
                ostatok = 0
                if p == 0:
                    if value[0][2].day > 14:
                        sum_MN = value[0][1] + standart_chislo * value[0][2].month
                        if sum_MN >= value[0][3]:
                            rez_MN = sum_MN - value[0][3]
                            dat_MN = value[0][2] + datetime.timedelta(days=rez_MN)
                            for row in sheet.iter_rows():
                                for cell1 in row:
                                    if cell1.value == key:
                                        for cell2 in row:
                                            if cell2.value == value[0][2]:
                                                cell1_row = cell1.row
                                                cell2_row = cell2.row
                                                if cell1_row == cell2_row:
                                                    sheet[end_ostatok + str(cell2.row)] = rez_MN
                                                    sheet[end_date + str(cell2.row)] = dat_MN
                                                    sheet[conclusion + str(cell2.row)] = 'Достаточно'
                        else:
                            rez_MN = sum_MN - value[0][3]
                            for row in sheet.iter_rows():
                                for cell1 in row:
                                    if cell1.value == key:
                                        for cell2 in row:
                                            if cell2.value == value[0][2]:
                                                cell1_row = cell1.row
                                                cell2_row = cell2.row
                                                if cell1_row == cell2_row:
                                                    sheet[end_ostatok + str(cell2.row)] = rez_MN
                                                    sheet[conclusion + str(cell2.row)] = 'Недостаточно'


                    else:
                        sum_MN = value[0][1] + standart_chislo * (value[0][2].month - 1)
                        if sum_MN >= value[0][3]:
                            rez_MN = sum_MN - value[0][3]
                            dat_MN = value[0][2] + datetime.timedelta(days=rez_MN)
                            for row in sheet.iter_rows():
                                for cell1 in row:
                                    if cell1.value == key:
                                        for cell2 in row:
                                            if cell2.value == value[0][2]:
                                                cell1_row = cell1.row
                                                cell2_row = cell2.row
                                                if cell1_row == cell2_row:
                                                    sheet[end_ostatok + str(cell2.row)] = rez_MN
                                                    sheet[end_date + str(cell2.row)] = dat_MN
                                                    sheet[conclusion + str(cell2.row)] = 'Достаточно'

                        else:
                            rez_MN = sum_MN - value[0][3]
                            for row in sheet.iter_rows():
                                for cell1 in row:
                                    if cell1.value == key:
                                        for cell2 in row:
                                            if cell2.value == value[0][2]:
                                                cell1_row = cell1.row
                                                cell2_row = cell2.row
                                                if cell1_row == cell2_row:
                                                    sheet[end_ostatok + str(cell2.row)] = rez_MN
                                                    sheet[conclusion + str(cell2.row)] = 'Недостаточно'



                else:
                    value[p][1] = rez_MN
                    if value[p-1][2].day >14:
                        if value[p][2].day > 14:
                            raznitca_of_the_month = value[p][2].month - value[p - 1][2].month
                            sum_MN = value[p][1] + standart_chislo * raznitca_of_the_month
                            if sum_MN >= value[p][3]:
                                rez_MN = sum_MN - value[p][3]
                                dat_MN = value[p][2] + datetime.timedelta(days=rez_MN)

                                for Row in sheet.iter_rows():
                                    for Cell1 in Row:
                                        if Cell1.value == key:
                                            for Cell2 in Row:
                                                if Cell2.value == value[p][2]:
                                                    Cell1_row = Cell1.row
                                                    Cell2_row = Cell2.row
                                                    if Cell1_row == Cell2_row:
                                                        sheet[end_ostatok + str(Cell2.row)] = rez_MN
                                                        sheet[end_date + str(Cell2.row)] = dat_MN
                                                        sheet[conclusion + str(Cell2.row)] = 'Достаточно'
                            else:
                                rez_MN = sum_MN - value[p][3]
                                for Row in sheet.iter_rows():
                                    for Cell1 in Row:
                                        if Cell1.value == key:
                                            for Cell2 in Row:
                                                if Cell2.value == value[p][2]:
                                                    Cell1_row = Cell1.row
                                                    Cell2_row = Cell2.row
                                                    if Cell1_row == Cell2_row:
                                                        sheet[end_ostatok + str(Cell2.row)] = rez_MN
                                                        sheet[conclusion + str(Cell2.row)] = 'Недостаточно'



                        else:
                            raznitca_of_the_month = value[p][2].month - value[p - 1][2].month
                            sum_MN = value[p][1] + standart_chislo * (raznitca_of_the_month - 1)
                            if sum_MN >= value[p][3]:
                                rez_MN = sum_MN - value[p][3]
                                dat_MN = value[p][2] + datetime.timedelta(days=rez_MN)
                                for Row in sheet.iter_rows():
                                    for Cell1 in Row:
                                        if Cell1.value == key:
                                            for Cell2 in Row:
                                                if Cell2.value == value[p][2]:
                                                    Cell1_row = Cell1.row
                                                    Cell2_row = Cell2.row
                                                    if Cell1_row == Cell2_row:
                                                        sheet[end_ostatok + str(Cell2.row)] = rez_MN
                                                        sheet[end_date + str(Cell2.row)] = dat_MN
                                                        sheet[conclusion + str(Cell2.row)] = 'Достаточно'
                            else:
                                rez_MN = sum_MN - value[p][3]
                                for Row in sheet.iter_rows():
                                    for Cell1 in Row:
                                        if Cell1.value == key:
                                            for Cell2 in Row:
                                                if Cell2.value == value[p][2]:
                                                    Cell1_row = Cell1.row
                                                    Cell2_row = Cell2.row
                                                    if Cell1_row == Cell2_row:
                                                        sheet[end_ostatok + str(Cell2.row)] = rez_MN
                                                        sheet[conclusion + str(Cell2.row)] = 'Недостаточно'
                    else:
                        if value[p][2].day > 14:
                            raznitca_of_the_month = value[p][2].month - value[p - 1][2].month
                            sum_MN = value[p][1] + standart_chislo * (raznitca_of_the_month + 1)
                            if sum_MN >= value[p][3]:
                                rez_MN = sum_MN - value[p][3]
                                dat_MN = value[p][2] + datetime.timedelta(days=rez_MN)

                                for Row in sheet.iter_rows():
                                    for Cell1 in Row:
                                        if Cell1.value == key:
                                            for Cell2 in Row:
                                                if Cell2.value == value[p][2]:
                                                    Cell1_row = Cell1.row
                                                    Cell2_row = Cell2.row
                                                    if Cell1_row == Cell2_row:
                                                        sheet[end_ostatok + str(Cell2.row)] = rez_MN
                                                        sheet[end_date + str(Cell2.row)] = dat_MN
                                                        sheet[conclusion + str(Cell2.row)] = 'Достаточно'
                            else:
                                rez_MN = sum_MN - value[p][3]
                                for Row in sheet.iter_rows():
                                    for Cell1 in Row:
                                        if Cell1.value == key:
                                            for Cell2 in Row:
                                                if Cell2.value == value[p][2]:
                                                    Cell1_row = Cell1.row
                                                    Cell2_row = Cell2.row
                                                    if Cell1_row == Cell2_row:
                                                        sheet[end_ostatok + str(Cell2.row)] = rez_MN
                                                        sheet[conclusion + str(Cell2.row)] = 'Недостаточно'



                        else:
                            raznitca_of_the_month = value[p][2].month - value[p - 1][2].month
                            sum_MN = value[p][1] + standart_chislo * raznitca_of_the_month
                            if sum_MN >= value[p][3]:
                                rez_MN = sum_MN - value[p][3]
                                dat_MN = value[p][2] + datetime.timedelta(days=rez_MN)
                                for Row in sheet.iter_rows():
                                    for Cell1 in Row:
                                        if Cell1.value == key:
                                            for Cell2 in Row:
                                                if Cell2.value == value[p][2]:
                                                    Cell1_row = Cell1.row
                                                    Cell2_row = Cell2.row
                                                    if Cell1_row == Cell2_row:
                                                        sheet[end_ostatok + str(Cell2.row)] = rez_MN
                                                        sheet[end_date + str(Cell2.row)] = dat_MN
                                                        sheet[conclusion + str(Cell2.row)] = 'Достаточно'
                            else:
                                rez_MN = sum_MN - value[p][3]
                                for Row in sheet.iter_rows():
                                    for Cell1 in Row:
                                        if Cell1.value == key:
                                            for Cell2 in Row:
                                                if Cell2.value == value[p][2]:
                                                    Cell1_row = Cell1.row
                                                    Cell2_row = Cell2.row
                                                    if Cell1_row == Cell2_row:
                                                        sheet[end_ostatok + str(Cell2.row)] = rez_MN
                                                        sheet[conclusion + str(Cell2.row)] = 'Недостаточно'

        else:
            if value[0][2].day > 14:
                sum_OD = value[0][1] + standart_chislo * value[0][2].month
                if sum_OD >= value[0][3]:
                    rez_OD = sum_OD - value[0][3]
                    dat_OD = value[0][2] + datetime.timedelta(days=rez_OD)
                    for row in sheet.iter_rows():
                        for cell1 in row:
                            if cell1.value == key:
                                for cell2 in row:
                                    if cell2.value == value[0][2]:
                                        cell1_row = cell1.row
                                        cell2_row = cell2.row
                                        if cell1_row == cell2_row:
                                            sheet[end_ostatok + str(cell2.row)] = rez_OD
                                            sheet[end_date + str(cell2.row)] = dat_OD
                                            sheet[conclusion + str(cell2.row)] = 'Достаточно'
                else:
                    rez_OD = sum_OD - value[0][3]
                    for row in sheet.iter_rows():
                        for cell1 in row:
                            if cell1.value == key:
                                for cell2 in row:
                                    if cell2.value == value[0][2]:
                                        cell1_row = cell1.row
                                        cell2_row = cell2.row
                                        if cell1_row == cell2_row:
                                            sheet[end_ostatok + str(cell2.row)] = rez_OD
                                            sheet[conclusion + str(cell2.row)] = 'Недостаточно'


            else:
                sum_OD = value[0][1] + standart_chislo * (value[0][2].month - 1)
                if sum_OD >= value[0][3]:
                    rez_OD = sum_OD - value[0][3]
                    dat_OD = value[0][2] + datetime.timedelta(days=rez_OD)
                    for row in sheet.iter_rows():
                        for cell1 in row:
                            if cell1.value == key:
                                for cell2 in row:
                                    if cell2.value == value[0][2]:
                                        cell1_row = cell1.row
                                        cell2_row = cell2.row
                                        if cell1_row == cell2_row:
                                            sheet[end_ostatok + str(cell2.row)] = rez_OD
                                            sheet[end_date + str(cell2.row)] = dat_OD
                                            sheet[conclusion + str(cell2.row)] = 'Достаточно'

                else:
                    rez_OD = sum_OD - value[0][3]
                    for row in sheet.iter_rows():
                        for cell1 in row:
                            if cell1.value == key:
                                for cell2 in row:
                                    if cell2.value == value[0][2]:
                                        cell1_row = cell1.row
                                        cell2_row = cell2.row
                                        if cell1_row == cell2_row:
                                            sheet[end_ostatok + str(cell2.row)] = rez_OD
                                            sheet[conclusion + str(cell2.row)] = 'Недостаточно'



    book.save(url)



